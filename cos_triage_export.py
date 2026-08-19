"""Export active CoS loops to Excel triage spreadsheets — one per mailbox.

Usage:
    python cos_triage_export.py                 # one workbook per mailbox
    python cos_triage_export.py <output_path>   # single workbook, all mailboxes
    python cos_triage_export.py --mailbox cfm [<output_path>]

One mailbox = one workbook, so each inbox is triaged on its own. The mailbox list
lives in cos/mailboxes.py; loops that belong to no registered mailbox go to an
"Unattributed" workbook, and only when there are any.

Default output (timestamp shared across a batch, so the importer can find them
as a set):
    data/triage/CoS Triage YYYY-MM-DD HH-MM - <Mailbox>.xlsx

Columns (research-backed layout — GTD next-actions + Eisenhower matrix):
    #             Stable loop number
    Urgency       urgent / high / normal / low
    Dir           On You / Waiting / FYI / Deferred
    Action Type   Reply / Approve / Pay / Decide / Review / FYI
    Counterparty
    Summary
    Category
    Age           Days the loop has been open
    Due           Hard deadline (Excel date)
    Email Date    Original send date (Excel date)
    Sentiment     Only shown for concerned / frustrated / angry
    Link          Deep link back to Front
    Triage Action Fill in: done / drop / snooze / etc.
    Notes         Free text — saved on loop record when imported
    _id           Loop ID — hidden; used by importer
"""
import datetime
import os
import sys
from pathlib import Path

os.environ.setdefault("LEDGER_BACKEND", os.environ.get("LEDGER_BACKEND", "firestore"))
os.environ.setdefault("GCP_PROJECT", os.environ.get("GCP_PROJECT", "cfm-front-mail"))

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env", override=True)

from cos import ledger, mailboxes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colours — urgency × direction (Eisenhower-inspired) ──────────────────────
# Key: (urgency, direction) → ARGB hex
_URGENCY_FILL: dict[tuple[str, str], str] = {
    ("urgent",  "i_owe"):      "FFFFD0D0",  # bright red
    ("urgent",  "owed_to_me"): "FFFFEAD0",  # bright orange
    ("high",    "i_owe"):      "FFFFE4E4",  # light red
    ("high",    "owed_to_me"): "FFFFF4E0",  # light orange
    ("normal",  "i_owe"):      "FFFFF0F0",  # pale pink
    ("normal",  "owed_to_me"): "FFFFFFF0",  # pale yellow
    ("low",     "i_owe"):      "FFF8F8F8",  # near-white
    ("low",     "owed_to_me"): "FFFAFAFA",  # near-white
}
_FYI_FILL      = "FFF5F5F5"  # grey
_DEFERRED_FILL = "FFE3F2FD"  # blue
_DEFAULT_FILL  = "FFFAFAFA"

HEADER  = "FF2C5F8A"   # navy header
DIVIDER = "FFD0D8E4"   # section divider

THIN   = Side(style="thin", color="FFD0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATE_FMT = "MM/DD/YYYY"

# ── Column definitions ────────────────────────────────────────────────────────
COLS = [
    ("#",             5),
    ("Urgency",       9),
    ("Dir",           9),
    ("Action Type",  12),
    ("Counterparty", 22),
    ("Email",        30),   # sender's address — needed to tell us what to exclude
    ("Summary",      50),
    ("Category",     13),
    ("Age",           6),
    ("Due",          11),
    ("Email Date",   11),
    ("Sentiment",    11),
    ("Link",         10),
    ("Triage Action",22),
    ("Notes",        35),
    ("_id",           0),   # hidden — do not delete; used by importer
]

# 1-based column index lookup — derived from COLS so additions never break these
_COL = {name: idx for idx, (name, _) in enumerate(COLS, 1)}

# Urgency sort order
_URGENCY_ORDER = {"urgent": 0, "high": 1, "normal": 2, "low": 3}


def _local_now() -> datetime.datetime:
    tz_name = os.environ.get("COS_TIMEZONE", "America/New_York")
    try:
        import zoneinfo
        tz = zoneinfo.ZoneInfo(tz_name)
        return datetime.datetime.now(tz)
    except Exception:
        return datetime.datetime.now()


def _local_today() -> datetime.date:
    return _local_now().date()


def _parse_date(s: str | None) -> datetime.date | None:
    if not s:
        return None
    try:
        return datetime.date.fromisoformat(str(s)[:10])
    except (ValueError, TypeError):
        return None


def _age_days(first_seen: str | None, today: datetime.date) -> int | None:
    d = _parse_date(first_seen)
    return (today - d).days if d else None


def _triage_action_list() -> str:
    """The Action dropdown, with delegation targets pulled from the importer's
    own DELEGATES map so the sheet can never offer an action the importer does
    not understand (or miss one it does)."""
    base = ["done", "drop", "exclude", "subscribe", "fyi", "defer"]
    try:
        from cos_triage_import import DELEGATES
        base += list(DELEGATES)
    except Exception:
        base += ["delegate to admin", "send to sally"]
    base += ["snooze 1d", "snooze 3d", "snooze 1w", "snooze 2w", "snooze 1m"]
    joined = ",".join(base)
    # Excel caps an inline validation list at 255 chars and fails silently past
    # it — the dropdown just stops appearing. Adding a few more delegates would
    # reach that; say so loudly rather than shipping a sheet with no dropdown.
    if len(joined) > 250:
        print(f"  WARNING: Action dropdown is {len(joined)} chars (Excel limit 255). "
              f"Move the list to a hidden sheet and reference it by range.")
    return joined


def _row_fill(loop: dict, *, deferred: bool = False) -> str:
    if deferred:
        return _DEFERRED_FILL
    if loop.get("fyi"):
        return _FYI_FILL
    urgency   = (loop.get("urgency") or "normal").lower()
    direction = loop.get("direction", "owed_to_me")
    return _URGENCY_FILL.get((urgency, direction), _DEFAULT_FILL)


def _sort_key(loop: dict) -> tuple:
    if loop.get("fyi"):
        return (10, 0, 0, loop.get("first_seen") or "")
    urgency   = (loop.get("urgency") or "normal").lower()
    urg_order = _URGENCY_ORDER.get(urgency, 4)
    dir_order = 0 if loop["direction"] == "i_owe" else 1
    # Oldest first within same urgency+direction — longest-waiting surfaces first
    age_key   = loop.get("first_seen") or "9999"
    return (urg_order, dir_order, age_key, 0)


def export(output_path: str | None = None, *, mailbox: str = "",
           timestamp: str = "") -> str:
    """Write one triage workbook. mailbox="" means every loop, unsplit."""
    now       = _local_now()
    today     = now.date()
    today_str = today.isoformat()
    mb_label  = mailboxes.label_for(mailbox) if mailbox else "All mailboxes"

    if not output_path:
        out_dir = Path(__file__).parent / "data" / "triage"
        out_dir.mkdir(parents=True, exist_ok=True)
        # Include time so multiple same-day exports don't overwrite each other
        # and so a locked file (Excel open) generates a fresh name automatically.
        # The timestamp is passed in by export_all() so every workbook in one
        # batch shares it — that's what lets the importer treat them as a set.
        # Date stays immediately after "CoS Triage " so the importer's existing
        # date glob still matches.
        ts = timestamp or now.strftime("%Y-%m-%d %H-%M")
        suffix = f" - {mailboxes.slug(mailbox)}" if mailbox else ""
        output_path = str(out_dir / f"CoS Triage {ts}{suffix}.xlsx")

    loops    = ledger.list_loops(mailbox=mailbox)
    deferred = ledger.list_loops(deferred_only=True, mailbox=mailbox)

    loops.sort(key=_sort_key)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Triage"
    ws.freeze_panes = "A2"

    # ── Header row ───────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color="FFFFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor=HEADER)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for col_idx, (name, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    # ── Row writer ────────────────────────────────────────────────────────────
    def _write_loop_row(ws, row_idx: int, loop: dict, *, is_deferred: bool = False):
        fill = PatternFill("solid", fgColor=_row_fill(loop, deferred=is_deferred))

        urgency  = (loop.get("urgency") or "normal").lower()
        dir_raw  = loop.get("direction", "owed_to_me")

        if is_deferred:
            dir_label = "Deferred"
        elif loop.get("fyi"):
            dir_label = "FYI"
        elif dir_raw == "i_owe":
            dir_label = "On You"
        else:
            dir_label = "Waiting"

        # Sentiment: only surface negative values
        raw_sentiment = (loop.get("sentiment") or "").lower()
        sentiment_display = raw_sentiment if raw_sentiment in (
            "concerned", "frustrated", "angry") else ""

        age = _age_days(loop.get("first_seen"), today)

        values = [
            loop.get("num") or "",           # #
            urgency,                          # Urgency
            dir_label,                        # Dir
            loop.get("action_type") or "",    # Action Type
            loop.get("counterparty") or "",       # Counterparty
            loop.get("counterparty_email") or "",  # Email
            loop.get("summary") or "",             # Summary
            loop.get("category") or "",       # Category
            age if age is not None else "",   # Age
            _parse_date(loop.get("due_at")),  # Due (Excel date)
            _parse_date(loop.get("source_date")),  # Email Date (Excel date)
            sentiment_display,                # Sentiment
            loop.get("source_link") or "",    # Link
            "",                               # Triage Action — user fills
            "",                               # Notes — user fills
            loop.get("id") or "",             # _id (hidden)
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill   = fill
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_idx == _COL["Summary"]),
            )

            # Date cells — native Excel date format
            if col_idx in (_COL["Due"], _COL["Email Date"]) and val is not None:
                cell.number_format = DATE_FMT

            # Summary row height
            if col_idx == _COL["Summary"]:
                ws.row_dimensions[row_idx].height = 40

            # Age color: red ≥30d, orange ≥14d
            if col_idx == _COL["Age"] and isinstance(val, int):
                if val >= 30:
                    cell.font = Font(bold=True, color="FFCC0000")
                elif val >= 14:
                    cell.font = Font(bold=True, color="FFCC6600")

            # Urgency label bold for urgent/high
            if col_idx == _COL["Urgency"] and urgency in ("urgent", "high"):
                cell.font = Font(bold=True)

            # Link — hyperlink
            if col_idx == _COL["Link"] and val:
                cell.hyperlink = val
                cell.value = "open"
                cell.font  = Font(color="FF0563C1", underline="single")

    # ── Main rows ─────────────────────────────────────────────────────────────
    next_row = 2
    for loop in loops:
        _write_loop_row(ws, next_row, loop)
        next_row += 1

    # ── Deferred section ──────────────────────────────────────────────────────
    if deferred:
        div_fill = PatternFill("solid", fgColor=DIVIDER)
        div_font = Font(bold=True, size=10, color="FF555555")
        label = (f"Deferred — Review Later  "
                 f"({len(deferred)} item{'s' if len(deferred) != 1 else ''})")
        for col_idx in range(1, len(COLS) + 1):
            cell = ws.cell(row=next_row, column=col_idx,
                           value=(label if col_idx == 1 else ""))
            cell.fill   = div_fill
            cell.font   = div_font
            cell.border = BORDER
        ws.row_dimensions[next_row].height = 16
        next_row += 1

        for loop in deferred:
            _write_loop_row(ws, next_row, loop, is_deferred=True)
            next_row += 1

    # ── Triage Action dropdown ────────────────────────────────────────────────
    # Only when there is at least one data row: a mailbox with nothing open
    # yields total_rows == 1, and "M2:M1" is an invalid range openpyxl rejects.
    # (Latent before the mailbox split, when there was always ≥1 loop.)
    action_col  = _COL["Triage Action"]
    total_rows  = next_row - 1
    if total_rows >= 2:
        dv = DataValidation(
            type="list",
            formula1=f'"{_triage_action_list()}"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = (f"{get_column_letter(action_col)}2:"
                    f"{get_column_letter(action_col)}{total_rows}")
        ws.add_data_validation(dv)

    # ── Column widths + hide _id ──────────────────────────────────────────────
    for col_idx, (name, width) in enumerate(COLS, start=1):
        col_letter = get_column_letter(col_idx)
        if width == 0:
            ws.column_dimensions[col_letter].hidden = True
        else:
            ws.column_dimensions[col_letter].width = width

    # ── Auto-filter ───────────────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

    ws.sheet_properties.tabColor = "2C5F8A"

    # ── Instructions sheet ────────────────────────────────────────────────────
    info = wb.create_sheet("Instructions")
    instructions = [
        (f"CoS Triage Spreadsheet — {mb_label}", True),
        ("", False),
        (f"Generated: {today_str}  |  "
         f"{len(loops)} active + {len(deferred)} deferred loops", False),
        (f"Mailbox: {mb_label}"
         + (f"  <{mailboxes.address_for(mailbox)}>" if mailbox and
            mailboxes.address_for(mailbox) else "")
         + ("  — this workbook covers ONLY this mailbox." if mailbox else ""), False),
        ("", False),
        ("COLOUR KEY:", True),
        ("  Rows are coloured by urgency × direction:", False),
        ("  Bright red    = Urgent / On You", False),
        ("  Bright orange = Urgent / Waiting", False),
        ("  Light red     = High / On You", False),
        ("  Light orange  = High / Waiting", False),
        ("  Pale pink     = Normal / On You", False),
        ("  Pale yellow   = Normal / Waiting", False),
        ("  Grey          = FYI (auto-clears in 24h)", False),
        ("  Blue          = Deferred (parked for later)", False),
        ("", False),
        ("AGE COLUMN:", True),
        ("  Orange bold = open 14+ days.  Red bold = open 30+ days.", False),
        ("", False),
        ("EMAIL COLUMN:", True),
        ("  The sender's actual address. If a sender should never appear here at", False),
        ("  all, say so and it gets a permanent exclude rule — no AI cost, no row.", False),
        ("  Use the Sender Rules sheet to add one yourself: action = exclude.", False),
        ("", False),
        ("TRIAGE ACTIONS:", True),
        ("  done          — mark the loop as completed", False),
        ("  drop          — discard (won't appear again)", False),
        ("  exclude       — drop + tag as junk (trains future classification)", False),
        ("  subscribe     — tag in Front as reading-list + drop", False),
        ("  fyi           — re-classify as notification; auto-clears in 24h", False),
        ("  defer         — move to Deferred section; hidden from briefing", False),
        ("  delegate to admin — email it to admin@cfmins.org; they own it now", False),
        ("  send to sally    — email it to Sally Swygert; she owns it now", False),
        ("  snooze 1d/3d/1w/2w/1m  — hide until that time", False),
        ("  snooze YYYY-MM-DD      — hide until specific date", False),
        ("  (blank)       — no action; Notes still saved if filled in", False),
        ("", False),
        ("DEFERRED SECTION (blue rows at bottom):", True),
        ("  Use done/drop/snooze on a deferred row to fully resolve it.", False),
        ("", False),
        ("NOTES:", True),
        ("  Fill in the Notes column on any row — saved for any action, "
         "or even with no action.", False),
        ("", False),
        ("IMPORTANT:", True),
        ("  DO NOT delete or edit the _id column — the importer uses it.", False),
        ("  Save file, then run: Run CoS Triage.bat → option 2 (Import)", False),
    ]
    for text, bold in instructions:
        info.append([text])
        if bold:
            info.cell(info.max_row, 1).font = Font(bold=True, size=12)
    info.column_dimensions["A"].width = 72

    # ── Sender Rules sheet (FILTER-1 / PRIORITY-1) ───────────────────────────
    _write_sender_rules_sheet(wb)

    # ── Guidance sheet (GUIDANCE-1) ───────────────────────────────────────────
    _write_guidance_sheet(wb)

    try:
        wb.save(output_path)
    except PermissionError:
        print(f"\nERROR: Cannot write to:\n  {output_path}\n"
              "The file may be open in Excel. Close it and try again.")
        sys.exit(1)
    print(f"Exported {len(loops)} active + {len(deferred)} deferred "
          f"[{mb_label}] -> {output_path}")
    return output_path


def export_all(*, include_empty: bool = True) -> list[tuple[str, str]]:
    """Write one workbook per mailbox. Returns [(mailbox_key, path), ...].

    include_empty keeps a registered mailbox's workbook even with zero loops, so
    the morning email always has the same shape. The Unattributed bucket is the
    exception — it only appears when something actually landed in it, since an
    empty one is just noise.
    """
    ts = _local_now().strftime("%Y-%m-%d %H-%M")
    written: list[tuple[str, str]] = []
    for mb in mailboxes.mailboxes(include_unassigned=True):
        key = mb["key"]
        n = len(ledger.list_loops(mailbox=key))
        if n == 0 and (key == mailboxes.UNASSIGNED or not include_empty):
            continue
        written.append((key, export(mailbox=key, timestamp=ts)))
    return written


def _write_sender_rules_sheet(wb: openpyxl.Workbook) -> None:
    from cos import ledger as _ledger
    rules = _ledger.list_sender_rules()

    ws = wb.create_sheet("Sender Rules")
    ws.sheet_properties.tabColor = "C05A00"

    HDR_COLS = [
        ("Email / Domain", 28),
        ("Action",         14),
        ("Category",       14),
        ("Direction",      13),
        ("Importance",     11),
        ("Subject Pattern",22),
        ("Notes",          35),
        ("_delete",         8),  # user puts 'yes' to remove the rule on import
    ]
    hdr_font  = Font(bold=True, color="FFFFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="FFC05A00")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for ci, (name, width) in enumerate(HDR_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    row_fill = PatternFill("solid", fgColor="FFFFF8F0")
    for ri, rule in enumerate(rules, 2):
        vals = [
            rule.get("email") or "",
            rule.get("action") or "",
            rule.get("category") or "",
            rule.get("direction") or "",
            rule.get("importance") or "",
            rule.get("subject_pattern") or "",
            rule.get("notes") or "",
            "",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = row_fill; cell.border = BORDER
            cell.alignment = Alignment(vertical="top")

    # Action dropdown
    dv = DataValidation(type="list",
                        formula1='"exclude,fyi,force-category,subscribe"',
                        allow_blank=True, showDropDown=False)
    if len(rules) > 0:
        dv.sqref = f"B2:B{len(rules)+1}"
        ws.add_data_validation(dv)

    # Instructions in col I
    instrs = [
        "SENDER RULES — pre-classify senders before Claude is invoked.",
        "",
        "Action values:",
        "  exclude        — skip entirely (no loop created)",
        "  fyi            — force FYI classification",
        "  force-category — create loop but override category",
        "  subscribe      — tag as reading-list and drop",
        "",
        "Email / Domain:",
        "  Exact:  user@example.com",
        "  Domain: @example.com  (matches all addresses at that domain)",
        "  Subdomain: @hq.bill.com  (matches hq.bill.com only)",
        "",
        "Importance: 1-5 overrides AI urgency (PRIORITY-1).",
        "Subject Pattern: regex matched against summary.",
        "_delete: put 'yes' to remove a rule on next import.",
        "",
        "Save and run: Run CoS Triage.bat -> option 2 (Import)",
    ]
    ws.column_dimensions["I"].width = 55
    for ri, line in enumerate(instrs, 1):
        cell = ws.cell(row=ri, column=9, value=line)
        if ri == 1:
            cell.font = Font(bold=True)


def _write_guidance_sheet(wb: openpyxl.Workbook) -> None:
    from cos import ledger as _ledger
    items = _ledger.list_guidance()

    ws = wb.create_sheet("Guidance")
    ws.sheet_properties.tabColor = "1A6B3A"

    HDR_COLS = [
        ("Key",    18),
        ("Scope",  20),
        ("Body",   65),
        ("Active",  8),
        ("_delete", 8),
    ]
    hdr_font  = Font(bold=True, color="FFFFFFFF", size=11)
    hdr_fill  = PatternFill("solid", fgColor="FF1A6B3A")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for ci, (name, width) in enumerate(HDR_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    row_fill = PatternFill("solid", fgColor="FFF0FFF4")
    for ri, g in enumerate(items, 2):
        vals = [
            g.get("key") or "",
            g.get("scope") or "all",
            g.get("body") or "",
            "yes" if g.get("active") else "no",
            "",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = row_fill; cell.border = BORDER
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(ci == 3))  # Body wraps
        ws.row_dimensions[ri].height = 40

    # Active dropdown
    dv_active = DataValidation(type="list", formula1='"yes,no"',
                               allow_blank=True, showDropDown=False)
    if items:
        dv_active.sqref = f"D2:D{len(items)+1}"
        ws.add_data_validation(dv_active)

    # Instructions in col F
    instrs = [
        "GUIDANCE — standing instructions injected into every AI analysis.",
        "",
        "Key: short slug (e.g. wire-confirmations)",
        "Scope: all | category:finance | sender:atlanticunionbank.com",
        "Body: plain English instruction for Claude.",
        "Active: yes/no  — toggle without deleting.",
        "_delete: put 'yes' to remove on next import.",
        "",
        "Examples:",
        '  "Wire transfer confirmation emails are always FYI."',
        '  "Parish emails asking about payment status are i_owe / finance / high."',
        "",
        "Save and run: Run CoS Triage.bat -> option 2 (Import)",
    ]
    ws.column_dimensions["F"].width = 58
    for ri, line in enumerate(instrs, 1):
        cell = ws.cell(row=ri, column=6, value=line)
        if ri == 1:
            cell.font = Font(bold=True)


if __name__ == "__main__":
    args = sys.argv[1:]
    mailbox = ""
    if "--mailbox" in args:
        i = args.index("--mailbox")
        mailbox = args[i + 1] if i + 1 < len(args) else ""
        del args[i:i + 2]
        if mailbox and not mailboxes.by_key(mailbox):
            sys.exit(f"Unknown mailbox {mailbox!r}. "
                     f"Known: {', '.join(mailboxes.keys(include_unassigned=True))}")
    path = args[0] if args else None
    if path or mailbox:
        export(path, mailbox=mailbox)
    else:
        for key, p in export_all():
            print(f"  {key:10s} {p}")
