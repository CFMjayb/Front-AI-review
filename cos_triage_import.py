"""Import a reviewed CoS triage spreadsheet and apply actions to Firestore.

Usage:
    python cos_triage_import.py [path_to_xlsx]

If path_to_xlsx is omitted, picks the most recent file in data/triage/.

Action column values (case-insensitive):
    done                  — resolve loop as done
    drop                  — resolve loop as dropped
    exclude               — drop + tag as junk (helps CoS avoid re-classifying similar emails)
    subscribe             — tag in Front as "cos/reading-list" + drop; view latest in Front
    fyi                   — re-classify as FYI notification (grey, auto-clears 24h)
    defer                 — move to Deferred section; hidden from main list and briefing
    snooze YYYY-MM-DD     — snooze until that date
    snooze 1d / 2d / 3d   — snooze for N days
    snooze 1w / 2w        — snooze for N weeks
    snooze 1m             — snooze for 1 month
    (blank)               — no action change; Notes column still saved if filled in
"""
import datetime
import os
import re
import sys
from pathlib import Path
from typing import Any, Optional

os.environ.setdefault("LEDGER_BACKEND", os.environ.get("LEDGER_BACKEND", "firestore"))
os.environ.setdefault("GCP_PROJECT", os.environ.get("GCP_PROJECT", "cfm-front-mail"))

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env", override=True)

from cos import ledger
import openpyxl

# ── Lazy Front client (created on first archive call) ────────────────────────
_front_client: Optional[Any] = None


def _get_front() -> Any:
    global _front_client
    if _front_client is None:
        from front_client import FrontClient
        from auth import get_front_api_token
        _front_client = FrontClient(get_front_api_token())
    return _front_client


def _archive_in_front(loop_rec: Optional[dict], num: Any) -> bool:
    """Archive the source Front conversation when a loop is resolved.

    Thin wrapper over cos/front_archive.py, the single implementation every
    caller (this importer, the retirement scripts, the pipeline's exclude
    paths) shares — so "taken off the triage list" and "archived in Front"
    can never drift apart again the way they did on 2026-08-18.
    """
    from cos import front_archive
    return front_archive.archive_loop(_get_front(), loop_rec, printer=print)


def _parse_snooze_until(value: str) -> str | None:
    """Parse snooze target. Returns ISO datetime string or None on failure."""
    v = value.strip().lower()
    now = datetime.datetime.now(datetime.timezone.utc)

    # snooze YYYY-MM-DD
    m = re.match(r"snooze\s+(\d{4}-\d{2}-\d{2})$", v)
    if m:
        return m.group(1) + "T00:00:00Z"

    # snooze Nd (days) or Nw (weeks) or Nm (months)
    m = re.match(r"snooze\s+(\d+)([dwm])$", v)
    if m:
        n, unit = int(m.group(1)), m.group(2)
        if unit == "d":
            delta = datetime.timedelta(days=n)
        elif unit == "w":
            delta = datetime.timedelta(weeks=n)
        else:  # months — approximate as 30 days each
            delta = datetime.timedelta(days=n * 30)
        until = now + delta
        return until.strftime("%Y-%m-%dT%H:%M:%SZ")

    return None


_TS_RE = re.compile(r"CoS Triage (\d{4}-\d{2}-\d{2}(?: \d{2}-\d{2})?)")


# ── Delegation ───────────────────────────────────────────────────────────────
# Triage action -> (recipient, human name). Handing an item to someone emails
# them the loop and takes it off Jay's list; they now own it.
#
# Add a person here and they become a new dropdown option automatically — the
# export builds the Triage Action list from these keys, so the sheet and the
# importer can never drift apart.
DELEGATES: dict[str, tuple[str, str]] = {
    "delegate to admin": (os.environ.get("COS_ADMIN_EMAIL", "admin@cfmins.org"),
                          "the admin mailbox"),
    "send to sally":     (os.environ.get("COS_SALLY_EMAIL",
                                         "sswygert@episcopalmaryland.org"),
                          "Sally Swygert"),
}


def _delegate(loop: dict, num, note: str, to_email: str, to_name: str) -> bool:
    """Email a loop to someone else so they can carry it.

    Returns False on any failure. The caller then leaves the loop OPEN — a
    delegation that did not actually send must never silently clear the item,
    or the task disappears with nobody holding it.
    """
    if not loop:
        return False
    try:
        from cos import sender
    except Exception as exc:
        print(f"      sender unavailable: {exc}")
        return False

    cp = loop.get("counterparty") or "unknown"
    cp_email = loop.get("counterparty_email") or ""
    subject = f"[Delegated] #{num} - {cp}: {(loop.get('summary') or '')[:70]}"
    body = [
        f"Jay has delegated this open item to {to_name} to complete.",
        "",
        f"**From:** {cp}" + (f" <{cp_email}>" if cp_email else ""),
        f"**Item:** {loop.get('summary') or '(no summary)'}",
    ]
    if loop.get("category"):
        body.append(f"**Category:** {loop['category']}")
    if loop.get("due_at"):
        body.append(f"**Due:** {loop['due_at'][:10]}")
    if loop.get("source_link"):
        body.append(f"**Open in Front:** [{loop['source_link']}]({loop['source_link']})")
    if note:
        body += ["", f"**Jay's note:** {note}"]
    body += ["", "-- Chief of Staff"]

    try:
        sender.send(subject=subject, body_md="\n".join(body), to=[to_email])
        return True
    except Exception as exc:
        print(f"      delegation send failed: {exc}")
        return False


def _find_latest_exports() -> list[Path]:
    """Every workbook in the most recent export batch, newest batch first.

    The export writes one workbook per mailbox — "CoS Triage <ts> - <Mailbox>.xlsx"
    — with a shared timestamp. All files carrying the newest timestamp are one
    batch and must all be imported, or a mailbox's edits are silently discarded.
    Older single-file exports (no mailbox suffix) still match and still work.
    """
    triage_dir = Path(__file__).parent / "data" / "triage"
    if not triage_dir.exists():
        return []
    dated = sorted(
        triage_dir.glob("CoS Triage [0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]*.xlsx"),
        reverse=True)
    dated = [p for p in dated if not p.name.startswith("~$")]  # Excel lock files
    if not dated:
        return []
    stamped = [(m.group(1), p) for p in dated if (m := _TS_RE.match(p.name))]
    if not stamped:
        return [dated[0]]
    newest_ts = stamped[0][0]          # dated is already newest-first
    return sorted(p for ts, p in stamped if ts == newest_ts)


def _find_latest_export() -> Path | None:
    """Back-compat single-file accessor."""
    batch = _find_latest_exports()
    return batch[0] if batch else None


def _col_index(headers: list[str], name: str) -> int | None:
    try:
        return headers.index(name)
    except ValueError:
        return None


def _run_triage_sheet(wb: openpyxl.Workbook) -> dict:
    """Apply actions from the Triage sheet of an already-loaded workbook."""
    ws = wb["Triage"]

    # Read headers from row 1
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    idx_id     = _col_index(headers, "_id")
    idx_num    = _col_index(headers, "#")
    idx_action = _col_index(headers, "Triage Action") or _col_index(headers, "Action")
    idx_notes  = _col_index(headers, "Notes")

    if idx_id is None or idx_action is None:
        raise ValueError("Required columns '_id' or 'Action' not found. Was the file modified?")

    done = dropped = snoozed = skipped = errored = 0
    results = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        loop_id = str(row[idx_id] or "").strip()
        action  = str(row[idx_action] or "").strip().lower()
        notes   = str(row[idx_notes] or "").strip() if idx_notes is not None else ""
        num     = row[idx_num] if idx_num is not None else "?"

        if not loop_id:
            skipped += 1
            continue

        try:
            # Save notes first (for any action, including blank)
            if notes:
                existing = ledger.get_loop(loop_id)
                if existing:
                    combined = ((existing.get("notes") or "") + f"\n{notes}").strip()
                    ledger.patch_loop(loop_id, notes=combined)

            if not action:
                if notes:
                    print(f"  #{num} note saved")
                skipped += 1
                continue

            if action == "done":
                loop_rec = ledger.get_loop(loop_id)
                ledger.resolve_loop(loop_id, "done")
                if _archive_in_front(loop_rec, num):
                    ledger.patch_loop(loop_id, front_archived=True)
                print(f"  #{num} done")
                done += 1

            elif action == "drop":
                loop_rec = ledger.get_loop(loop_id)
                ledger.resolve_loop(loop_id, "dropped")
                if _archive_in_front(loop_rec, num):
                    ledger.patch_loop(loop_id, front_archived=True)
                print(f"  #{num} dropped")
                dropped += 1

            elif action == "exclude":
                loop_rec = ledger.get_loop(loop_id)
                ledger.patch_loop(loop_id, category="junk")
                ledger.resolve_loop(loop_id, "dropped", reason="excluded:junk")
                if _archive_in_front(loop_rec, num):
                    ledger.patch_loop(loop_id, front_archived=True)
                print(f"  #{num} excluded (junk)")
                dropped += 1

            elif action in DELEGATES:
                # Hand the item to someone else and take it off Jay's list.
                to_email, to_name = DELEGATES[action]
                loop_rec = ledger.get_loop(loop_id)
                if _delegate(loop_rec, num, notes, to_email, to_name):
                    ledger.resolve_loop(loop_id, "done", reason=f"delegated:{to_email}")
                    # Jay's own copy is spoken for now — archive it like any other
                    # resolved loop, same as done/drop/exclude below.
                    if _archive_in_front(loop_rec, num):
                        ledger.patch_loop(loop_id, front_archived=True)
                    print(f"  #{num} delegated to {to_name} <{to_email}>")
                    done += 1
                else:
                    # The send failed — leave the loop OPEN. Resolving it here
                    # would drop the item on the floor with nobody holding it.
                    print(f"  #{num} WARNING: delegation email failed - loop left open")
                    errored += 1

            elif action == "subscribe":
                loop_rec = ledger.get_loop(loop_id)
                if loop_rec and loop_rec.get("channel") == "front":
                    try:
                        _get_front().add_tag(loop_rec["source_ref"], "cos/reading-list")
                        print(f"  #{num} tagged in Front as cos/reading-list")
                    except Exception as exc:
                        print(f"  #{num} WARNING: could not tag in Front: {exc}")
                ledger.resolve_loop(loop_id, "dropped", reason="subscribed:reading-list")
                print(f"  #{num} subscribed")
                dropped += 1

            elif action == "fyi":
                ledger.patch_loop(loop_id, fyi=True, deferred=False)
                print(f"  #{num} marked FYI")
                done += 1

            elif action == "defer":
                ledger.patch_loop(loop_id, deferred=True)
                print(f"  #{num} deferred")
                skipped += 1  # not resolved — just moved to deferred section

            elif action.startswith("snooze"):
                until = _parse_snooze_until(action)
                if not until:
                    print(f"  #{num} ERROR: can't parse snooze date from '{action}'")
                    errored += 1
                    continue
                ledger.snooze_loop(loop_id, until)
                print(f"  #{num} snoozed until {until[:10]}")
                snoozed += 1

            else:
                print(f"  #{num} unknown action '{action}' — skipped")
                skipped += 1

        except Exception as exc:
            print(f"  #{num} ERROR: {exc}")
            errored += 1

    print(f"\nTriage: Done: {done}  Dropped: {dropped}  Snoozed: {snoozed}  "
          f"Skipped: {skipped}  Errors: {errored}")
    return {
        "done": done, "dropped": dropped, "snoozed": snoozed,
        "skipped": skipped, "errored": errored,
        "total_actioned": done + dropped + snoozed,
    }


def _import_sender_rules(wb: openpyxl.Workbook) -> dict:
    if "Sender Rules" not in wb.sheetnames:
        return {"upserted": 0, "deleted": 0}
    ws = wb["Sender Rules"]
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    def _ci(name):
        try: return headers.index(name)
        except ValueError: return None

    idx_email   = _ci("Email / Domain")
    idx_action  = _ci("Action")
    idx_cat     = _ci("Category")
    idx_dir     = _ci("Direction")
    idx_imp     = _ci("Importance")
    idx_subj    = _ci("Subject Pattern")
    idx_notes   = _ci("Notes")
    idx_del     = _ci("_delete")

    if idx_email is None or idx_action is None:
        print("  Sender Rules sheet: missing required columns — skipped")
        return {"upserted": 0, "deleted": 0}

    upserted = deleted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        email = str(row[idx_email] or "").strip().lower()
        if not email:
            continue
        delete_flag = str(row[idx_del] or "").strip().lower() == "yes" if idx_del is not None else False
        if delete_flag:
            if ledger.delete_sender_rule(email):
                print(f"  sender-rule deleted: {email}")
                deleted += 1
            continue
        action = str(row[idx_action] or "").strip().lower()
        if not action:
            continue
        imp_raw = row[idx_imp] if idx_imp is not None else None
        try:
            imp = int(float(str(imp_raw))) if imp_raw else 0
        except (ValueError, TypeError):
            imp = 0
        ledger.upsert_sender_rule(
            email=email, action=action,
            category=str(row[idx_cat] or "").strip() if idx_cat is not None else "",
            direction=str(row[idx_dir] or "").strip() if idx_dir is not None else "",
            importance=imp,
            subject_pattern=str(row[idx_subj] or "").strip() if idx_subj is not None else "",
            notes=str(row[idx_notes] or "").strip() if idx_notes is not None else "",
        )
        print(f"  sender-rule upserted: {email} -> {action}")
        upserted += 1

    return {"upserted": upserted, "deleted": deleted}


def _import_guidance(wb: openpyxl.Workbook) -> dict:
    if "Guidance" not in wb.sheetnames:
        return {"upserted": 0, "deleted": 0}
    ws = wb["Guidance"]
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]

    def _ci(name):
        try: return headers.index(name)
        except ValueError: return None

    idx_key    = _ci("Key")
    idx_scope  = _ci("Scope")
    idx_body   = _ci("Body")
    idx_active = _ci("Active")
    idx_del    = _ci("_delete")

    if idx_key is None or idx_body is None:
        print("  Guidance sheet: missing required columns — skipped")
        return {"upserted": 0, "deleted": 0}

    upserted = deleted = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = str(row[idx_key] or "").strip().lower()
        if not key:
            continue
        delete_flag = str(row[idx_del] or "").strip().lower() == "yes" if idx_del is not None else False
        if delete_flag:
            if ledger.delete_guidance(key):
                print(f"  guidance deleted: {key}")
                deleted += 1
            continue
        body = str(row[idx_body] or "").strip()
        if not body:
            continue
        scope  = str(row[idx_scope] or "all").strip() if idx_scope is not None else "all"
        active_raw = str(row[idx_active] or "yes").strip().lower() if idx_active is not None else "yes"
        active = active_raw not in ("no", "false", "0")
        ledger.upsert_guidance(key=key, body=body, scope=scope or "all", active=active)
        print(f"  guidance upserted: {key} (active={active})")
        upserted += 1

    return {"upserted": upserted, "deleted": deleted}


def run_import(xlsx_path: str | None = None) -> dict:
    # Re-named old run_import to _run_triage_import; wrapper calls both
    return _run_all_imports(xlsx_path)


def _run_all_imports(xlsx_path: str | None = None) -> dict:
    """Import one file, or — given no path — every file in the latest batch."""
    if not xlsx_path:
        batch = _find_latest_exports()
        if not batch:
            raise FileNotFoundError("No triage file found in data/triage/. Run export first.")
        if len(batch) > 1:
            print(f"Latest batch has {len(batch)} mailbox workbooks:")
            for b in batch:
                print(f"  - {b.name}")
            totals = {"done": 0, "dropped": 0, "snoozed": 0}
            per_file = []
            for b in batch:
                r = _run_one_file(str(b))
                per_file.append({"file": b.name, **r})
                for k in totals:
                    totals[k] += r.get(k, 0) or 0
            remaining = len(ledger.list_loops())
            print(f"\nBatch totals: Done: {totals['done']}  "
                  f"Dropped: {totals['dropped']}  Snoozed: {totals['snoozed']}")
            print(f"Active loops remaining in Firestore: {remaining}")
            return {**totals, "files": per_file}
        xlsx_path = str(batch[0])

    return _run_one_file(xlsx_path)


def _run_one_file(xlsx_path: str) -> dict:
    print(f"Reading: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path)

    # ── Triage sheet ─────────────────────────────────────────────────────────
    triage_result = _run_triage_sheet(wb)

    # ── Sender Rules sheet ────────────────────────────────────────────────────
    sr = _import_sender_rules(wb)
    if sr["upserted"] or sr["deleted"]:
        print(f"\nSender Rules: {sr['upserted']} upserted, {sr['deleted']} deleted")

    # ── Guidance sheet ────────────────────────────────────────────────────────
    gd = _import_guidance(wb)
    if gd["upserted"] or gd["deleted"]:
        print(f"Guidance: {gd['upserted']} upserted, {gd['deleted']} deleted")

    remaining = len(ledger.list_loops())
    print(f"Active loops remaining in Firestore: {remaining}")
    return {**triage_result, "sender_rules": sr, "guidance": gd}


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else None
    _run_all_imports(path)
