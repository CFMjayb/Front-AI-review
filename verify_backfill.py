"""Verify backfill results and spot-check triage export."""
import os
from pathlib import Path
from collections import Counter

os.environ.setdefault("LEDGER_BACKEND", "firestore")
os.environ.setdefault("GCP_PROJECT", "cfm-front-mail")
os.environ.setdefault("USE_SECRET_MANAGER", "true")

from dotenv import load_dotenv
load_dotenv(Path(__file__).parent / ".env", override=True)

from cos import ledger

loops = ledger.list_loops()
total = len(loops)

has_sd      = sum(1 for l in loops if l.get("source_date"))
missing_sd  = [l for l in loops if not l.get("source_date")]
has_sent    = sum(1 for l in loops if l.get("sentiment"))
sent_vals   = Counter(l.get("sentiment") or "null" for l in loops)

print(f"=== Field Population ===")
print(f"Total active loops  : {total}")
print(f"source_date         : {has_sd}/{total}  ({len(missing_sd)} missing)")
print(f"sentiment           : {has_sent}/{total}")
print(f"sentiment breakdown : {dict(sorted(sent_vals.items()))}")
print()

print("=== 5-row sample (loops with source_date) ===")
samples = [l for l in loops if l.get("source_date")][:5]
for s in samples:
    num      = s.get("num", "?")
    party    = (s.get("counterparty") or "")[:22]
    urgency  = s.get("urgency", "")
    action   = s.get("action_type", "")
    sd       = (s.get("source_date") or "")[:10]
    sent     = s.get("sentiment") or ""
    print(f"  #{num:4}  {party:22}  urgency={urgency:6}  action={action:8}  "
          f"source_date={sd}  sentiment={sent}")

print()
if missing_sd:
    print(f"=== Loops still missing source_date ({len(missing_sd)}) ===")
    for l in missing_sd[:15]:
        print(f"  #{l.get('num'):4}  {l.get('source_ref','')}")

# ── Verify the export file ────────────────────────────────────────────────────
print()
print("=== Spot-checking export file ===")
import openpyxl, datetime
from cos_triage_export import COLS, _COL, export

test_path = "data/triage/CoS Triage VERIFY.xlsx"
export(test_path)

wb = openpyxl.load_workbook(test_path)
ws = wb.active
headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print(f"Headers ({len(headers)}): {headers}")
print()

# Check 5 data rows
print("First 5 data rows:")
for row_idx in range(2, 7):
    row = {headers[c]: ws.cell(row_idx, c + 1).value for c in range(len(headers))}
    print(f"  Row {row_idx}: urgency={row.get('Urgency')}  action={row.get('Action Type')}  "
          f"email_date={row.get('Email Date')}  sentiment={row.get('Sentiment')}  "
          f"age={row.get('Age')}  dir={row.get('Dir')}")

# Count non-blank values in new columns
blank_check = {}
for col_name in ("Urgency", "Dir", "Action Type", "Age", "Email Date", "Sentiment"):
    col_idx = _COL.get(col_name)
    if not col_idx:
        continue
    vals = [ws.cell(r, col_idx).value for r in range(2, ws.max_row + 1)
            if ws.cell(r, col_idx).value not in (None, "")]
    blank_check[col_name] = len(vals)

print()
print("Non-blank value counts per column:")
for col, count in blank_check.items():
    print(f"  {col:15}: {count}")

# Verify Email Date cells are actual date objects (not strings)
sd_col = _COL["Email Date"]
date_types = Counter(
    type(ws.cell(r, sd_col).value).__name__
    for r in range(2, min(ws.max_row + 1, 20))
    if ws.cell(r, sd_col).value is not None
)
print(f"\nEmail Date cell types (first 18 rows): {dict(date_types)}")
